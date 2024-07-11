"""文件、IO处理"""
import time

"""读取键盘输入
raw_input() : # Python 2.x 示例

input() :  Python 3.x中
在Python 3.x中，input() 函数等同于Python 2.x中的raw_input()函数，即它始终返回用户输入的字符串，不再尝试执行或解析输入内容。
    这样设计是为了避免安全问题和混淆，因为让程序直接执行未经检查的用户输入可能会导致严重的安全漏洞。
"""
# name = input("请输入您的名字：")
# print("您好，" + name)

"""打开和关闭文件
open() 函数是用于文件操作的基础函数，它允许你打开一个文件并返回一个文件对象，以便进行读、写或追加等操作
    open(file, mode='r', buffering=-1, encoding=None, errors=None, newline=None, closefd=True, opener=None)
    
file：必需参数，指定了要打开的文件路径，可以是相对路径或绝对路径。
mode：可选参数，指明打开文件的方式。常见的模式包括：
'x'：创建模式，仅在文件不存在的情况下创建并打开文件，若文件已存在则会抛出异常。
't'：文本模式（默认），与'r', 'w', 'a' 或 'x' 结合使用，处理文本文件。
'b'：二进制模式，与'r', 'w', 'a' 或 'x' 结合使用，处理二进制文件。

'+'：更新模式，结合其他模式如 'r+', 'w+', 'a+'，允许在同一文件上进行读写操作。

'r'（默认）：以只读方式打开文件。文件的指针将会放在文件的开头。这是默认模式。
'rb': 以二进制格式打开一个文件用于只读。文件指针将会放在文件的开头。这是默认模式。一般用于非文本文件如图片等。
'r+': 打开一个文件用于读写。文件指针将会放在文件的开头。
'rb+': 以二进制格式打开一个文件用于读写。文件指针将会放在文件的开头。一般用于非文本文件如图片等。

'w'：打开一个文件只用于写入。如果该文件已存在则打开文件，并从开头开始编辑，即原有内容会被删除。如果该文件不存在，创建新文件。
'wb+': 以二进制格式打开一个文件用于读写。如果该文件已存在则打开文件，并从开头开始编辑，即原有内容会被删除。如果该文件不存在，创建新文件。
    一般用于非文本文件如图片等。

'a'：打开一个文件用于追加。如果该文件已存在，文件指针将会放在文件的结尾。也就是说，新的内容将会被写入到已有内容之后。如果该文件不存在，创建新文件进行写入。
'ab': 以二进制格式打开一个文件用于追加。如果该文件已存在，文件指针将会放在文件的结尾。也就是说，新的内容将会被写入到已有内容之后。如果该文件不存在，创建新文件进行写入。
'a+': 打开一个文件用于读写。如果该文件已存在，文件指针将会放在文件的结尾。文件打开时会是追加模式。如果该文件不存在，创建新文件用于读写。
'ab+': 以二进制格式打开一个文件用于追加。如果该文件已存在，文件指针将会放在文件的结尾。如果该文件不存在，创建新文件用于读写。



buffering：设置缓冲策略，如果不指定，默认为 -1 表示系统默认缓冲。0表示不缓冲，1表示行缓冲，大于1的整数代表缓冲区大小。
encoding：如果文件是以文本模式打开的，可以指定字符编码（例如 'utf-8'）。
errors：在解码时指定如何处理错误，比如 'strict'（默认，引发错误）、'ignore'（忽略错误）或 'replace'（用特殊符号替换错误字符）。
newline：控制行结尾的处理方式，特别是跨平台处理换行符时有用。
closefd：布尔值，是否在关闭文件对象时也关闭底层的描述符，默认为 True。
opener：一个可调用对象，用来定义自定义的打开文件的方法，通常情况下不需要设置。

在完成文件操作后，必须确保调用文件对象的 close() 方法来关闭文件，以释放操作系统资源。现代编程实践中推荐使用上下文管理器（如 with 语句）来自动关闭文件：
     close() 方法刷新缓冲区里任何还没写入的信息，并关闭该文件，这之后便不能再进行写入。
"""

fo = open('foo.txt', 'w')
fo.close()
print(fo.name, fo.mode, fo.closed, fo.encoding)

"""wite()
write()方法可将任何字符串写入一个打开的文件。需要重点注意的是，Python字符串可以是二进制数据，而不是仅仅是文字。

write()方法不会在字符串的结尾添加换行符('\n')：
"""
fo = open("foo.txt", "w")
fo.write("www.runoob.com!\nVery good site!\n")

# 关闭打开的文件
fo.close()

"""read()方法
read（）方法从一个打开的文件中读取一个字符串。需要重点注意的是，Python字符串可以是二进制数据，而不是仅仅是文字。
    fileObject.read([count])
    count：可选参数，指定要读取的字节数或字符数（取决于文件是二进制模式还是文本模式）。
        如果省略了这个参数，或者传入了负值，read() 方法将尝试读取文件的剩余所有内容。
使用示例和行为：
"""
with open('foo.txt', 'r') as f:
    content = f.read()
print(content)  # 输出文件的全部内容

with open('foo.txt', 'r') as f:
    chunk = f.read(10)  # 读取前10个字符
print(chunk)

with open('foo.txt', 'r') as f:
    remaining_content = f.read(-1)  # 等价于 f.read()
print(remaining_content)

"""文件定位指的是改变文件对象内部的读写指针位置，以便从文件的不同位置开始读取或写入数据。这一操作通过 fileObject.seek(offset, whence) 方法实现。

fileObject.seek(offset, whence) 方法参数说明：

offset：这是要移动的字节数偏移量。如果该值为正数，则指针向文件末尾方向移动；如果是负数，则向文件开头方向移动。

whence（可选，默认值为0）：
    0（os.SEEK_SET 或 io.SEEK_SET）：表示相对于文件起始处的偏移量。
    1（os.SEEK_CUR 或 io.SEEK_CUR）：表示相对于当前文件指针位置的偏移量。
    2（os.SEEK_END 或 io.SEEK_END）：表示相对于文件结束处的偏移量。
"""
import os


def read_file_bytes(filename):
    """
    读取并处理文件的某些字节。

    :param filename: 要处理的文件名
    """
    # 检查文件是否存在且可读写
    if not os.path.exists(filename) or not os.access(filename, os.R_OK | os.W_OK):
        print(f"文件 {filename} 不存在或不可读写。")
        return

    try:
        # 打开文件
        with open(filename, 'r+b') as f:
            # 移动到文件开头
            f.seek(0)
            print("当前位置：", f.tell())

            # 读取第一个字符并打印
            first_char = f.read(1)
            print(first_char, end='')

            # 移动到当前位置之后5个字节的位置
            f.seek(5, 1)
            print("当前位置：", f.tell())

            # 读取新的位置上的一个字节
            next_char = f.read(1)
            print(next_char, end='')

            # 获取文件长度，以确保相对位置移动不会超出边界
            file_length = os.path.getsize(filename)

            # 移动到文件结尾前3个字节的位置，确保不会超出文件边界
            seek_position = max(0, file_length - 3)
            f.seek(seek_position, 2)
            print("当前位置：", f.tell())

            # 再次读取一个字节
            last_chars = f.read(3)
            print(last_chars)
    except IOError as e:
        print("文件操作失败:", e)
        # 确保在发生异常时正确处理资源
        # 在本例中，由于使用了with语句，文件会自动关闭
        # 但在更复杂的情况下，可能需要显式关闭资源


# 调用函数进行测试
read_file_bytes('foo.txt')

print('-------------------fileObject.flush() ----------------')
"""fileObject.flush() 是 Python 中用于文件操作的一个方法，它作用于打开的文件对象（例如通过 open() 函数创建的对象）。
调用此方法的主要目的是将缓冲区中的数据强制写入磁盘。
    在进行文件写入操作时，Python 会采用内部缓冲机制来优化性能。这意味着当你使用 fileObject.write() 方法写入数据时，并不会立即把数据同步到硬盘上，
而是先存放在内存缓冲区中，当缓冲区满或者程序结束时再一起写入硬盘。
    如果你需要确保所有已写入的数据立即保存到硬盘上，可以调用 fileObject.flush() 方法。这样做可以降低数据丢失的风险，尤其是在处理关键数据或日志记录等场景下。
"""
with open('foo.txt', 'w') as f:
    for i in range(20):
        f.write(f'Line {i}\n')  # 数据被写入缓冲区而非立即保存到磁盘
        f.flush()  # 强制将缓冲区的内容写入磁盘

print('----------------file.fileno()-----------------')
"""
file.fileno() 是 Python 内置的文件对象方法，主要用于获取与该文件对象关联的底层操作系统文件描述符（File Descriptor）。
    在Unix-like系统中，每个打开的文件或设备都会分配一个非负整数作为其文件描述符。这个描述符是内核用于标识进程中的打开文件的一种方式。
当你使用 open() 函数打开一个文件时，Python 会在后台为该文件创建一个文件对象，并将其与一个底层的操作系统文件描述符关联起来。
    通过调用 fileno() 方法，你可以获取到这个关联的文件描述符。
为什么需要 file.fileno()？
    进行更底层的系统调用：某些情况下，你需要对文件执行一些高级或特定的IO操作，如设置文件的读写位置、读取/修改文件状态等，
        这时可以使用os模块提供的如os.read()、os.write()、fcntl、mmap等函数，这些函数通常需要直接操作文件描述符。
    与其他库配合使用：有些第三方库可能要求以文件描述符形式传入文件句柄，而非Python文件对象。
        例如，在多路复用I/O（如select, epoll）或网络编程中，有时会需要将标准输入输出流或者普通文件的描述符传递给这些库。
"""

with open('foo.txt', 'r+b') as f:
    # 获取文件描述符
    file_descriptor = f.fileno()
    print(file_descriptor)
    # 使用os级别的read函数
    data = os.read(file_descriptor, 10)

    # 或者与其它需要文件描述符的API配合使用
    # 这里仅作示例，具体取决于你使用的API
    # with mmap.mmap(file_descriptor, length=0, access=mmap.ACCESS_READ) as mmapped_file:
# ... 对内存映射文件进行操作 ...

print('----------------file.isatty()-----------------')
"""file.isatty() 是 Python 文件对象的一个方法，主要用于检测该文件对象是否连接到一个终端设备（tty, terminal 或 console）。
    在操作系统中，TTY是Teletype Device的缩写，通常指代字符设备接口，用于与用户进行交互式输入输出。
isatty() 方法详解：
定义和功能： file.isatty() 方法无参数，直接调用即可。当调用此方法时，它会检查当前文件对象关联的流是否直接连接到一个终端或控制台。
    如果是，则返回 True；如果不是，即如果文件对象是一个普通文件、管道、网络套接字或者其他非终端设备，那么返回 False。

用途：
    在程序中，isatty() 通常用来决定是否应该显示彩色输出、进度条或者其它交互式内容，这些内容只有在终端环境下才有意义。
    当你希望程序的行为根据其是否直接与用户交互而改变时，可以使用这个方法来判断。
    例如，在脚本中，你可能不希望在重定向输出到文件时还打印出彩色日志，此时可以用 if sys.stdout.isatty(): 判断是否直接输出到终端。
"""
import sys

# 检查标准输入（stdin）是否连接到终端
if sys.stdin.isatty():
    print("您正在通过终端直接输入")
else:
    print("您的输入不是来自终端，可能是重定向或管道")

# 或者对打开的文件做同样的检查
with open('foo.txt', 'r') as f:
    if f.isatty():
        print("文件'some_file.txt' 直接连接到了一个终端设备")
    else:
        print("文件'some_file.txt' 不是一个终端设备")

print('----------------file.next()-----------------')
"""然而，在Python 3中，由于迭代器协议的变化，已经移除了file.next()方法。
    取而代之的是使用内置的next()函数配合文件对象进行迭代，或者直接遍历文件对象来逐行读取文件。
"""
# Python 3.x
with open('foo.txt', 'r') as f:
    # 使用内置的next()函数
    try:
        while True:
            line = next(f)
            print(line)
    except StopIteration:
        # 文件读取完毕时抛出StopIteration异常
        pass

# 或者更简洁地通过for循环遍历文件
with open('foo.txt', 'r') as f:
    for line in f:
        print(line)

print('----------------file.read([size])-----------------')
"""file.read([size]) 是 Python 中文件对象的一个方法，用于从文件中读取数据。该方法可以指定可选参数 size 来控制要读取的数据量。
详解：
基本用法：
    不带参数调用 file.read() 时，它会尝试读取并返回文件中的所有剩余内容，直到文件结束（EOF）为止。
    带有整数参数 size 调用 file.read(size)，则会尝试读取最多 size 字节的数据并返回。如果到达文件末尾，则返回已读到的数据，可能少于请求的 size 字节数。
返回值：
    返回类型通常是字符串（在Python 2中）或字节（在Python 3中）。在Python 3中，如果你希望读取文本文件，
    需要先使用 open(filename, 'r', encoding='encoding_type') 指定编码方式打开文件，然后调用 read() 方法将返回解码后的字符串。
"""
# Python 3 示例
with open('foo.txt', 'r', encoding='utf-8') as f:
    # 读取整个文件内容
    whole_content = f.read()
    print(whole_content)

    # 读取前5个字符
    f.seek(0)  # 重置读写指针位置
    first_five_chars = f.read(5)
    print(first_five_chars)

print('----------------file.readline([size])-----------------')
"""file.readline([size]) 是 Python 中文件对象的一个方法，用于从文件中读取一行内容。
详解：
基本用法：
    不带参数调用 file.readline() 时，它会读取并返回文件中的下一行，包括行尾的换行符（\n）。
    带有整数参数 size 调用 file.readline(size)，则最多读取包含 size 字节的一行数据，并返回。
        如果在到达指定字节数前遇到换行符，则会返回该行内容（包含换行符），即使实际读取的字节数小于 size。
返回值：
    返回类型通常是字符串（在Python 2中）或字节（在Python 3中处理二进制文件时）。
    对于文本文件，在Python 3中需要先使用正确的编码方式打开文件，如 open(filename, 'r', encoding='encoding_type')，
    然后 readline() 方法将返回解码后的单行字符串。
注意点：
    和 read() 方法一样，每次调用 readline() 后，文件的读取位置会自动前进到下一行的开始位置。
    如果最后一行没有换行符，调用 readline() 也会读取这一行，并返回其内容，但不包含任何额外的换行符。
    当读取到文件末尾时，返回一个空字符串 ''，表示文件已经结束。
"""
# Python 3 示例
with open('foo.txt', 'r', encoding='utf-8') as f:
    # 读取第一行内容
    first_line = f.readline()
    print(first_line)

    # 限制读取长度为10个字符的一行内容
    limited_length_line = f.readline(3)
    print(limited_length_line)

print('----------------file.readlines([sizeint])-----------------')
"""file.readlines([sizehint]) 是 Python 中文件对象的一个方法，用于从文件中读取所有行内容，并以列表形式返回。
详解：
基本用法：
    不带参数调用 file.readlines() 时，它会读取文件的每一行（包括换行符），并将每一行作为一个字符串元素放入列表中。最后返回包含所有行的列表。
    带有整数参数 sizehint 调用 file.readlines(sizehint)，这个参数是可选的，并且在大多数情况下其作用并不明显。
        Python 文档指出该参数的作用仅作为优化提示，表示预期的最大行数或最大总字节数，但实际上实现可能因平台而异。
        通常不推荐使用此参数，因为它可能导致不可预测的结果。
返回值：
    返回一个由字符串（在Python 2中）或字节（在处理二进制文件的Python 3中）组成的列表。
    对于文本文件，在Python 3中需要先指定正确的编码方式打开文件，如 open(filename, 'r', encoding='encoding_type')，
    然后 readlines() 方法将返回解码后的字符串列表。
    
使用 readlines() 方法可以方便地一次性获取文件中的所有行，但请注意如果文件非常大，这种方法可能会消耗大量内存。
    在这种情况下，考虑逐行读取或者分块读取数据更合适。
注：和 readline()、read() 方法一样，调用 readlines() 后，文件读取位置会移动到文件末尾。
"""
# Python 3 示例
with open('foo.txt', 'r', encoding='utf-8') as f:
    lines = f.readlines()
    print(type(lines))  # <class 'list'>
    for line in lines:
        print(line.strip())

print('----------------file.tell()-----------------')
"""用于获取当前文件读写指针的位置。这个位置通常以字节为单位表示，它指示了下一次读写操作将在文件中的起始位置。

该方法没有参数，调用后会返回一个整数，代表当前文件对象指向的文件内部偏移量。当你对文件进行读取或写入操作时，文件指针会根据读写的字节数自动更新。
功能与应用场景：
    当你想要知道在进行一系列文件操作后，当前文件读写位置在哪里时，可以使用 tell() 方法。
    在需要多次定位到文件的不同位置进行读写操作的情况下，先使用 tell() 获取当前位置，然后使用 seek(offset[, whence]) 方法移动文件指针到新的位置。
"""

with open('foo.txt', 'r+b') as f:
    # 获取文件打开时的初始位置
    initial_position = f.tell()
    print(f"初始位置: {initial_position} 字节")

    # 读取5个字节的数据
    data = f.read(5)
    print("读取的数据:", data)

    # 获取读取后的当前位置
    after_read_position = f.tell()
    print(f"读取后的位置: {after_read_position} 字节")

print('----------------file.truncate([size])-----------------')
"""file.truncate([size]) 是 Python 中文件对象的一个方法，用于截断或改变文件的大小。
详解：
基本用法：
    当调用 file.truncate() 时，如果未提供参数 size，则文件会被截断到当前文件指针的位置。这意味着从当前指针位置到文件末尾的所有内容都将被删除。
    如果提供了整数参数 size，则无论当前文件指针指向何处，文件都会被截断至指定的字节数。如果 size 小于当前文件大小，则文件将被缩短；
        如果 size 大于当前文件大小，并且操作系统支持扩展文件大小（大多数现代系统都支持），则文件会扩大并填充额外的空间，新扩展部分的内容通常是不确定的。

注意事项：
    在调用 truncate() 方法之前，请确保文件已以读写模式 'r+'、写入模式 'w+' 或追加模式 'a+' 打开，否则将引发异常。
    文件截断操作不会影响文件指针的位置，截断后，文件指针仍保持在原位置。
    对于只读文件或已经关闭的文件，调用 truncate() 方法会导致错误。
实际应用：
    该方法常用于清理文件的剩余内容，或者在知道文件需要特定大小的情况下，提前将其大小调整到位
"""
with open('foo.txt', 'r+') as f:
    # 移动文件指针到第10个字节
    f.seek(10)

    # 截断文件到当前指针位置，即保留前10个字节，删除其余部分
    f.truncate()

# 或者：

with open('foo.txt', 'r+') as f:
    # 截断文件到指定大小，例如截断为20字节大小
    f.truncate(20)

print('----------------file.write(str)-----------------')
"""这个方法接收一个字符串参数 str，并将其内容写入到当前文件对象所关联的文件中。
    它会从文件当前位置开始写入数据，并且不会在字符串末尾添加任何换行符或其他字符。
操作细节：
    如果文件是以 'w'（写入）、'a'（追加）或 'r+'、'w+'、'a+'（读写/追加模式）打开，则可以成功调用 write() 方法。
    调用 write() 后，文件指针会自动移动到下一个位置，即写入结束的位置。
    write() 方法返回的是写入的字节数，这有助于跟踪实际写入文件中的数据量。

"""
with open('example.txt', 'w') as f:
    # 写入一行文本
    f.write("Hello, World!")

# 或者写入多行文本
text = "Line 1\nLine 2\nLine 3\n"
with open('example.txt', 'a') as f:
    f.write(text)

"""
注意事项：
    如果需要在字符串之间添加换行，请确保在字符串中包含换行符 (\n)。
    在使用 'a' 模式（追加模式）时，每次调用 write() 都会在文件末尾追加内容，而不是覆盖现有内容。
    当以二进制模式打开文件时（如 'wb'），应传入字节串而非普通字符串到 write() 方法。例如：
"""
with open('binary_file.bin', 'wb') as f:
    f.write(b'Some binary data')

print('----------------file.writelines(sequence)-----------------')
"""file.writelines(sequence) 是 Python 中文件对象的一个方法，用于将一个字符串序列写入到已打开的文件中。
    这个方法通常用于一次性写入多个字符串或行内容到文件，并且每个字符串之间不会自动添加换行符。
注意事项：
    每个字符串元素都将按照它们在 sequence 中出现的顺序被写入到文件中。
    如果需要在字符串间自动插入换行符，需确保在序列中的每个字符串末尾都有换行符 \n。
    writelines() 方法不会在每两个字符串之间自动添加任何分隔符，即使在读取时看起来像是连续行也一样。
    和 write() 方法类似，调用 writelines() 后，文件指针会停留在最后一个写入位置。
    如果 sequence 包含非字符串类型的元素，将会引发 TypeError 异常。
错误处理：
    如果试图在只读文件或未正确打开的文件上调用 writelines() 方法，Python 将抛出异常。
"""
# 创建一个包含多行文本的列表
lines = ["Line 11\n", "Line 22\n", "Line 33"]

# 打开一个文件并追加内容（假设文件已存在或将以追加模式创建）
with open('example.txt', 'a') as f:
    # 使用 writelines 方法写入所有行
    f.writelines(lines)

# 注意：每一项字符串在被写入时都包含了其原有的换行符，所以这里会按预期形成多行文本

"""读取.xlsx文件"""
from openpyxl import load_workbook


# 打开一个 .xlsx 文件
# wb = load_workbook(filename='example.xlsx')

# 选择工作表（默认是第一个工作表，也可以通过名称或索引指定）
# sheet = wb.active  # 获取活动工作表
# 或者
# sheet = wb['Sheet1']  # 通过名称获取工作表

# 读取单元格内容
# cell_value = sheet['A1'].value  # 读取 A1 单元格的内容

# 遍历行和列来读取数据
# for row in sheet.iter_rows(values_only=True):  # 值模式下遍历所有行
#     for cell in row:
#         print(cell)  # 输出当前单元格的值

# 如果需要处理整个工作表的数据，可以将其转换为列表或其他数据结构
# data = []
# for row in sheet.iter_rows(values_only=True):
#     data.append([cell for cell in row])

# 关闭工作簿（虽然不是必须，但建议在完成操作后关闭以释放资源）
# wb.close()


class ExcelReader:
    """
    Excel文件读取器，支持上下文管理器。

    参数:
    - filename: 要读取的Excel文件的名称。

    方法:
    - __init__(self, filename): 构造函数，初始化Excel文件名称。
    - __enter__(self): 进入上下文时调用，打开Excel工作簿并返回活动工作表。
    - __exit__(self, exc_type, exc_val, exc_tb): 退出上下文时调用，关闭Excel工作簿。
    """

    def __init__(self, filename):
        """
        初始化ExcelReader实例。

        参数:
        - filename: 要读取的Excel文件的路径。
        """
        self.filename = filename

    def __enter__(self):
        """
        进入上下文管理器，打开Excel文件并返回活动工作表。

        返回:
        - 返回活动工作表以供后续操作。
        """
        self.wb = load_workbook(self.filename)  # 打开工作簿
        return self.wb.active  # 返回活动工作表

    def __exit__(self, exc_type, exc_val, exc_tb):
        """
        退出上下文管理器，关闭Excel工作簿。

        参数:
        - exc_type: 异常类型 (在异常发生时传递)。
        - exc_val: 异常值 (在异常发生时传递)。
        - exc_tb: 异常追踪 (在异常发生时传递)。
        """
        if not self.wb.is_closed:
            self.wb.close()  # 如果工作簿未关闭，则关闭它


# 使用示例：
# with ExcelReader('example.xlsx') as sheet:
#     cell_value = sheet['A1'].value  # 读取 A1 单元格的内容
#     print(cell_value)


# 上下文管理器会确保在完成操作后正确关闭工作簿，即使在处理过程中发生异常也是如此。


def read_lines_from_file(file_path):
    with open(file_path, 'r') as file:
        while line := file.readline():
            line = line.strip()
            if line:
                yield line


file_path = '/Users/chenhao/PycharmProjects/pythonProjectTest/99_pd_/abcde.txt'
for line in read_lines_from_file(file_path):
    print(line)
